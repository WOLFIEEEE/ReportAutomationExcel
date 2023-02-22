from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo

def table_creation_for_status(workbook, sheetname, df , last_row_num):

    for i, row in df.iterrows():
        if row['No_of_occurence'] > 0:
            df.at[i, 'Result'] = 'Fail'
    # Select the sheet by name

    new_df = df[['WCAG_SC', 'Result']];

    ws = workbook[sheetname]
    headers = list(new_df.columns);
    data = [row for row in [headers] + new_df.values.tolist()][1:]

    # for row_index, row in enumerate(data, last_row_num):
    #     for col_index, cell_value in enumerate(row, 1):
    #         cell = ws.cell(row=row_index, column=col_index, value=cell_value)
            
            # Set the color of the cell to light red if the value is 'Fail'
            # if cell_value == 'Fail':
            #     cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    # for i in col_index:
    #     data[0].append("=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=i, S=4, E=200))



    # Add the text at two rows below the table ended
    text_row_num = last_row_num + 2
    text_cell = ws.cell(row=text_row_num, column=2)
    text_cell.value = "Severity/Impact Wise Defect Distribution"
    text_cell.font = Font(size=13, color="2F5496")

    # Add the table one row below the text
    table_start_row = text_row_num + 2
    table_start_col = 2
    table_end_col = 3
    table_end_row = table_start_row + 50

    print(table_end_row)

    # # Set the headers and data for the table
    # headers = ["", "Critical", "High", "Medium", "Low"]
    # data = [["Defect count", "=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=C_column, S=4, E=200),
    #          "=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=H_column, S=4, E=200),
    #          "=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=M_column, S=4, E=200),
    #          "=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=L_column, S=4, E=200)]]

    # Add the table to the worksheet
    table_range = ws.cell(row=table_start_row, column=table_start_col).coordinate + ':' + ws.cell(row=table_end_row, column=table_end_col).coordinate
    tab = Table(displayName="Status_table", ref=table_range)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)

    # Write the headers and data to the table
    for i, header in enumerate(headers):
        cell = ws.cell(row=table_start_row, column=table_start_col+i)
        cell.value = header

    for row_idx, row_data in enumerate(data):
        for col_idx, col_data in enumerate(row_data):
            cell = ws.cell(row=table_start_row+1+row_idx, column=table_start_col+col_idx)
            cell.value = col_data

    cell_style_name = "cell_style"
    header_cell_style_name = "header_cell_style"
    for row in ws[table_range]:
        for cell in row:
            cell.style = cell_style_name

    for row in ws[table_range]:
        for cell in row:
            cell.style = header_cell_style_name
        break

    return tab.name
    # create_column_chart_issuetype(workbook, sheetname,"IssuetypeCountTable")
