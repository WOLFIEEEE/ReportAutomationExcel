from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from chart_creation.chart_creation_severity import create_column_chart_severity
from openpyxl.utils import column_index_from_string
from chart_creation.chart_creation_issuetype import create_column_chart_issuetype
def know_the_Columns(workbook, sheet_name, column_name):
    sheet = workbook[sheet_name]
    column_index = None

        # Search for the column named "WCAG SC#"
    for column in sheet.iter_cols():
        for cell in column:
            # print(cell.value)
            if cell.value == column_name:
                print(cell.column_letter)
                column_index = cell.column_letter

    sum_of_numbers = 0
    for i in range(4, 201):
        temp_sheet = workbook["Execution Summary"]
        column_index_temp = column_index_from_string(column_index)
        cell_value = temp_sheet.cell(row=i, column=column_index_temp).value
        if isinstance(cell_value, (int, float)):
            sum_of_numbers += cell_value
        if cell_value is None:
            break
    
    if(sum_of_numbers == 0):
        return 'XX';
        
    return column_index
def table_creation_for_issuetype(workbook, sheetname, last_row_num):
    # Select the sheet by name
    ws = workbook[sheetname]

    KN_column = know_the_Columns(workbook, "Execution Summary", "Keyboard Navigation")
    CC_column = know_the_Columns(workbook, "Execution Summary", "Color Contrast")
    C_column = know_the_Columns(workbook, "Execution Summary", "Color")
    Z_column = know_the_Columns(workbook, "Execution Summary", "Zoom")
    HTML_column = know_the_Columns(workbook, "Execution Summary", "HTML Validator")
    SR_column = know_the_Columns(workbook, "Execution Summary", "Screen Reader")
    O_column = know_the_Columns(workbook, "Execution Summary", "Other A11y")

    arr= [KN_column , CC_column , C_column , Z_column , HTML_column , SR_column , O_column]
    name_arr = [];
    col_index=[];
    for i in arr:
        temp_sheet = workbook["Execution Summary"]
        if(i == 'XX'):
            continue;
        column_index = column_index_from_string(i)
        cell_value = temp_sheet.cell(row=3, column=column_index).value
        name_arr.append(cell_value)
        col_index.append(i)
        print(cell_value)

    headers = [""] + name_arr;
    data = [["Defect count"]]

    for i in col_index:
        data[0].append("=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=i, S=4, E=200))



    # Add the text at two rows below the table ended
    text_row_num = last_row_num + 2
    text_cell = ws.cell(row=text_row_num, column=2)
    text_cell.value = "Severity/Impact Wise Defect Distribution"
    text_cell.font = Font(size=13, color="2F5496")

    # Add the table one row below the text
    table_start_row = text_row_num + 2
    table_start_col = 2
    table_end_col = headers.__len__() + 1
    table_end_row = table_start_row + 1

    print(table_end_row)

    # # Set the headers and data for the table
    # headers = ["", "Critical", "High", "Medium", "Low"]
    # data = [["Defect count", "=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=C_column, S=4, E=200),
    #          "=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=H_column, S=4, E=200),
    #          "=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=M_column, S=4, E=200),
    #          "=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=L_column, S=4, E=200)]]

    # Add the table to the worksheet
    table_range = ws.cell(row=table_start_row, column=table_start_col).coordinate + ':' + ws.cell(row=table_end_row, column=table_end_col).coordinate
    tab = Table(displayName="IssuetypeCountTable", ref=table_range)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)

    # Write the headers and data to the table
    for i, header in enumerate(headers):
        cell = ws.cell(row=table_start_row, column=table_start_col+i)
        cell.value = header

    for row_idx, row_data in enumerate(data):
        for col_idx, col_data in enumerate(row_data):
            cell = ws.cell(row=table_start_row+1+row_idx, column=table_start_col+col_idx)
            cell.value = col_data

    cell_style_name = "cell_style"
    for row in ws[table_range]:
        for cell in row:
            cell.style = cell_style_name
    create_column_chart_issuetype(workbook, sheetname,"IssuetypeCountTable")

    