from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle, Alignment, Font, PatternFill, Border, Side
from chart_creation.chart_creation_counts import create_column_chart
def add_dataframe_to_excel_sheet(workbook, sheetname, df):
    # create a new worksheet in the workbook
    sheet = workbook.create_sheet(sheetname)
    

    # add heading above the table
    # sheet.insert_rows(1)
    text_cell = sheet.cell(row=1, column=1, value="WCAG Rule Wise Defect Distribution Chart")
    text_cell.font = Font(size=13, color="2F5496")

    # add data from the dataframe to the sheet
    for r in dataframe_to_rows(df, index=False, header=True):
        sheet.append(r)

    # define the range of the table based on the size of the dataframe
    table_range = f"A2:B{len(df)+2}"

    # create a table from the range and add it to the sheet
    table = Table(displayName="WCAG_Counts", ref=table_range)

    # add style to the table
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)

    # set the header style
    header_style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False,
                                  showRowStripes=False, showColumnStripes=False)

    # set the style for the cells
    cell_style_name = "cell_style"
    if cell_style_name not in workbook.named_styles:
        cell_style = NamedStyle(name=cell_style_name)
        cell_style.alignment = Alignment(horizontal="left", vertical="center")
        cell_style.fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        cell_style.border = Border(left=Side(border_style='thin', color='000000'),
                                   right=Side(border_style='thin', color='000000'),
                                   top=Side(border_style='thin', color='000000'),
                                   bottom=Side(border_style='thin', color='000000'))
        workbook.add_named_style(cell_style)

    # set the style for the header cells
    header_cell_style_name = "header_cell_style"
    if header_cell_style_name not in workbook.named_styles:
        header_cell_style = NamedStyle(name=header_cell_style_name)
        header_cell_style.alignment = Alignment(horizontal="left", vertical="center")
        header_cell_style.fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
        header_cell_style.font = Font(bold=True)
        header_cell_style.border = Border(left=Side(border_style='thin', color='000000'),
                                          right=Side(border_style='thin', color='000000'),
                                          top=Side(border_style='thin', color='000000'),
                                          bottom=Side(border_style='thin', color='000000'))
        workbook.add_named_style(header_cell_style)

    # apply the styles to the cells in the table
    for row in sheet[table_range]:
        for cell in row:
            cell.style = cell_style_name

    # apply the style to the header cells
    for cell in sheet[2]:
        cell.style = header_cell_style_name

    table.tableStyleInfo = style
    sheet.add_table(table)
    
    sheet.insert_cols(1)
    sheet.insert_rows(1)
    sheet.insert_rows(3)
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    table_range = f"B4:C{len(df)+4}"
    table.ref = table_range
    create_column_chart(workbook, sheetname, "WCAG_Counts")
    return len(df)+5

