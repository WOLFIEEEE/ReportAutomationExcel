from openpyxl.styles import Font
from openpyxl.utils import range_boundaries
from openpyxl.worksheet.table import Table, TableStyleInfo
from chart_creation.chart_creation_conformacne import create_column_chart_conformance
def know_the_Columns(workbook, sheet_name, column_name):
    sheet = workbook[sheet_name]
    column_index = None

        # Search for the column named "WCAG SC#"
    for column in sheet.iter_cols():
        for cell in column:
            # print(cell.value)
            if cell.value == column_name:
                print(cell.column_letter)
                column_index = cell.column_letter

    return column_index
def table_creation_for_conformance(workbook, sheetname, last_row_num):
    # Select the sheet by name
    ws = workbook[sheetname]

    A_column = know_the_Columns(workbook, "Execution Summary", "Level A")
    AA_column = know_the_Columns(workbook, "Execution Summary", "Level AA")

    # Add the text at two rows below the table ended
    text_row_num = last_row_num + 2
    text_cell = ws.cell(row=text_row_num, column=2)
    text_cell.value = "Conformance Level Wise Defect Distribution"
    text_cell.font = Font(size=13, color="2F5496")

    # Add the table one row below the text
    table_start_row = text_row_num + 2
    table_start_col = 2
    table_end_col = 4
    table_end_row = table_start_row + 1

    print(table_end_row)

    # Set the headers and data for the table
    headers = ["Conformance Level", "A", "AA"]
    data = [["Defect count", "=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=A_column, S=4, E=200),
             "=SUM('Execution Summary'!{C}{S}:{C}{E} )".format(C=AA_column, S=4, E=200)]]

    # Add the table to the worksheet
    table_range = ws.cell(row=table_start_row, column=table_start_col).coordinate + ':' + ws.cell(row=table_end_row, column=table_end_col).coordinate
    tab = Table(displayName="ConformanceCountTable", ref=table_range)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    ws.add_table(tab)

    # Write the headers and data to the table
    for i, header in enumerate(headers):
        cell = ws.cell(row=table_start_row, column=table_start_col+i)
        cell.value = header

    for row_idx, row_data in enumerate(data):
        for col_idx, col_data in enumerate(row_data):
            cell = ws.cell(row=table_start_row+1+row_idx, column=table_start_col+col_idx)
            cell.value = col_data

    cell_style_name = "cell_style"
    for row in ws[table_range]:
        for cell in row:
            cell.style = cell_style_name
    create_column_chart_conformance(workbook, sheetname, "ConformanceCountTable")
    

    